import os
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def exportEmployeeData(records):
    if records:
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.append(['Id', 'Emri', 'Mbiemri', 'Pozicioni', 'Datëlindja','Telefoni', 'Email'])
            file_path = "C:\\Users\\" + os.getlogin() +  "\\Desktop\\Punonjesit.xlsx"

            sheet = workbook['Sheet']

            sheet.title = 'Prezencat'

            for row in records:
                ws.append(row)

            tab = Table(displayName="Table1", ref="A1:G"+str(len(records) + 1))

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

            tab.tableStyleInfo = style
            ws.add_table(tab)
            workbook.save(file_path)
            messagebox.showinfo("Mesazh"," Dokumenti u eksportua me sukses!")

        except FileNotFoundError:
            messagebox.showinfo("Mesazh", " Dokumenti nuk mund te eksportohej!")
    else:
        messagebox.showerror("Gabim!", "Tabela eshte bosh!")



def exportEmployeeActivityData(records):

    if records:
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.append(['Emri', 'Mbiemri','Pozicioni', 'Data', 'Ora Hyrje','Ora Dalje'])

            file_path = "C:\\Users\\" + os.getlogin() +  "\\Desktop\\HyrjetDaljet.xlsx"

            sheet = workbook['Sheet']

            sheet.title = 'Prezencat'

            for row in records:
                ws.append(row)

            tab = Table(displayName="Table1", ref="A1:G"+str(len(records) + 1))

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

            tab.tableStyleInfo = style
            ws.add_table(tab)
            workbook.save(file_path)
            messagebox.showinfo("Mesazh"," Dokumenti u eksportua me sukses!")

        except FileNotFoundError:
            messagebox.showinfo("Mesazh"," Dokumenti nuk mund te eksportohej!")
    else:
        messagebox.showerror("Gabim!", "Tabela eshte bosh!")


def exportEmployeWage(records):
    if records:
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.append(['Emri', 'Mbiemri', 'Pozicioni', 'Paga Orare'])
            file_path = "C:\\Users\\" + os.getlogin() + "\\Desktop\\Paga_Orare.xlsx"

            sheet = workbook['Sheet']

            sheet.title = 'Prezencat'

            for row in records:
                ws.append(row)

            tab = Table(displayName="Table1", ref="A1:G" + str(len(records) + 1))

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

            tab.tableStyleInfo = style
            ws.add_table(tab)
            workbook.save(file_path)
            messagebox.showinfo("Mesazh", " Dokumenti u eksportua me sukses!")
        except FileNotFoundError:
            messagebox.showinfo("Mesazh", " Dokumenti nuk mund te eksportohej!")
    else:
        messagebox.showerror("Gabim!", "Tabela eshte bosh!")



def exportEmployeHours(records):
    if records:
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.append(['Emri', 'Mbiemri', 'Data', 'Oret'])
            file_path = "C:\\Users\\" + os.getlogin() + "\\Desktop\\Oret.xlsx"

            sheet = workbook['Sheet']

            sheet.title = 'Prezencat'

            for row in records:
                ws.append(row)

            tab = Table(displayName="Table1", ref="A1:G" + str(len(records) + 1))

            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)

            tab.tableStyleInfo = style
            ws.add_table(tab)
            workbook.save(file_path)
            messagebox.showinfo("Mesazh", " Dokumenti u eksportua me sukses!")
        except FileNotFoundError:
            messagebox.showinfo("Mesazh", " Dokumenti nuk mund te eksportohej!")
    else:
        messagebox.showerror("Gabim!", "Tabela eshte bosh!")
